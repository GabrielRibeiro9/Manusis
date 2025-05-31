from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl
from collections import defaultdict
from datetime import datetime
from selenium.common.exceptions import TimeoutException
from datetime import timedelta
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import smtplib
from email.message import EmailMessage
import ssl
import os

def lancar_ordem_para_grupo(driver, log, num_os, linhas, header):
    lancamentos_realizados = 0
    wait = WebDriverWait(driver, 10)

    try:
        driver.switch_to.window(driver.window_handles[0])
        botao_abrir_ordem = wait.until(
            EC.element_to_be_clickable((By.XPATH, '//a[contains(@href, "abre_janela_apontaosplan")]'))
        )
        botao_abrir_ordem.click()
        WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
        driver.switch_to.window(driver.window_handles[-1])
        sleep(3)
        grupo_raw = linhas[0][header.index("GRUPO")]
        if grupo_raw is None or str(grupo_raw).strip() == "":
            raise ValueError(f"Grupo inválido detectado na OSP {num_os}, linhas: {len(linhas)}")
        grupo = str(grupo_raw).strip()
        log(f"Nova ordem para OSP {num_os} sendo aberta para grupo {grupo}.")
        lancamentos_realizados += 1

 # Ordena colaboradores pela data e hora de início
        linhas_ordenadas = sorted(
            linhas,
            key=lambda l: (
                l[header.index("ini_exec_real")] if isinstance(l[header.index("ini_exec_real")], datetime)
                else datetime.strptime(str(l[header.index("ini_exec_real")]), "%d/%m/%Y"),
                l[header.index("hor_ini_real")] if isinstance(l[header.index("hor_ini_real")], datetime)
                else datetime.strptime(str(l[header.index("hor_ini_real")]).strip() + ":00"
                                      if len(str(l[header.index("hor_ini_real")]).strip().split(':')) == 2
                                      else str(l[header.index("hor_ini_real")]), "%H:%M:%S")
            )
        )

        base = linhas_ordenadas[0]
        ini_exec_real = base[header.index("ini_exec_real")]
        hor_ini_real = base[header.index("hor_ini_real")]
        patrimonio = base[header.index("PATRIMONIO_ELECTROLUX")]
        re = base[header.index("RE")]
        des_servico = base[header.index("des_servico")]
        falha = str(base[header.index("FALHA")] or "").strip()
        defeito = str(base[header.index("DEFEITO")] or "").strip()
        solucao = str(base[header.index("SOLUCAO")] or "").strip()

        colaborador_base = base[header.index("COLABORADOR")]
        log(f"Base da OSP {num_os}, grupo {grupo}: colaborador {colaborador_base} - {ini_exec_real.strftime('%d/%m/%Y')} {hor_ini_real}")

        # Preenche DATA_PROG, DATA_ABRE, HORA_ABRE com o colaborador base
        data_prog = wait.until(EC.presence_of_element_located((By.ID, "DATA_PROG")))
        data_prog.click()
        data_prog.send_keys(Keys.CONTROL + "a")
        data_prog.send_keys(Keys.DELETE)
        data_prog.send_keys(ini_exec_real.strftime("%d/%m/%Y"))

        data_abre = wait.until(EC.presence_of_element_located((By.ID, "DATA_ABRE")))
        data_abre.click()
        data_abre.send_keys(Keys.CONTROL + "a")
        data_abre.send_keys(Keys.DELETE)
        data_abre.send_keys(ini_exec_real.strftime("%d/%m/%Y"))

        hora_abre = wait.until(EC.presence_of_element_located((By.ID, "cc[HORA_ABRE]")))
        hora_abre.click()
        hora_abre.send_keys(Keys.CONTROL + "a")
        hora_abre.send_keys(Keys.DELETE)
        hora_abre.send_keys(hor_ini_real.strftime("%H:%M:%S"))

        # Patrimonio e seleção da máquina
        patrimonio_input = driver.find_element(By.ID, "campo_filtro_cc[MID_MAQUINA]")
        patrimonio_input.clear()
        patrimonio_input.send_keys(str(patrimonio))
        log(f"Patrimônio encontrado: {patrimonio}")
        patrimonio_input.send_keys(Keys.ENTER)
        sleep(1)
        patrimonio_input.send_keys(Keys.ENTER)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//select[@id="cc[MID_MAQUINA]"]/option[not(@value="0")]'))
        )
        sleep(2)

        select_maquina = Select(driver.find_element(By.ID, "cc[MID_MAQUINA]"))
        select_maquina.select_by_index(1)
        sleep(2)

        driver.find_element(By.ID, "cc[SOLICITANTE]").send_keys("MANUFATURA")
        sleep(2)
        driver.find_element(By.ID, "cc[RESPONSAVEL]").send_keys(str(re))
        sleep(2)
        driver.find_element(By.ID, "cc[TIPO_SERVICO]").send_keys("PREVENTIVA")
        sleep(2)
        driver.find_element(By.ID, "cc[NATUREZA]").send_keys("3")
        sleep(2)
        driver.find_element(By.ID, "cc[MID_PRIORIDADE]").send_keys("NORMAL")
        sleep(2)
        driver.find_element(By.ID, "cc[TEXTO]").send_keys("PREVENTIVA")
        sleep(2)
        driver.find_element(By.ID, "cc[SOLUCAO_TEXTO]").send_keys(str(des_servico))
        sleep(2)

        def selecionar_opcao_parcial(select_element, texto_planilha):
            for option in select_element.options:
                if texto_planilha.strip().lower() in option.text.strip().lower():
                    select_element.select_by_visible_text(option.text)
                    return option.text
            raise ValueError(f"Opção '{texto_planilha}' não encontrada no dropdown.")

        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cc[CAUSA]")))
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cc[DEFEITO]")))
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "cc[SOLUCAO]")))

            select_falha = Select(driver.find_element(By.ID, "cc[CAUSA]"))
            sleep(1)
            select_defeito = Select(driver.find_element(By.ID, "cc[DEFEITO]"))
            sleep(1)
            select_solucao = Select(driver.find_element(By.ID, "cc[SOLUCAO]"))
            sleep(1)

            selecionar_opcao_parcial(select_falha, falha)
            sleep(1)
            selecionar_opcao_parcial(select_defeito, defeito)
            sleep(1)
            selecionar_opcao_parcial(select_solucao, solucao)
            sleep(1)

        except Exception as e:
            log(f"Erro ao selecionar opções do dropdown: {e}")

        sleep(0.5)

        # Lançar todos colaboradores do grupo
        for linha in linhas:
            status = linha[header.index("STATUS MANUSIS")]
            grupo_linha = str(linha[header.index("GRUPO")]).strip()
            if status == "PENDENTE" and grupo_linha == grupo:
                colaborador = linha[header.index("COLABORADOR")]
                ini_exec_real = linha[header.index("ini_exec_real")]
                hor_ini_real = linha[header.index("hor_ini_real")]
                fim_exec_real = linha[header.index("fim_exec_real")]

                driver.find_element(By.ID, "func").send_keys(str(colaborador))
                sleep(1)
                driver.find_element(By.ID, "fdatai").clear()
                driver.find_element(By.ID, "fdatai").send_keys(ini_exec_real.strftime("%d/%m/%Y"))
                sleep(1)
                driver.find_element(By.ID, "fhorai").clear()
                driver.find_element(By.ID, "fhorai").send_keys(f"{str(hor_ini_real)}:00" if len(str(hor_ini_real)) <= 5 else str(hor_ini_real))
                sleep(1)
                driver.find_element(By.ID, "fdataf").clear()
                driver.find_element(By.ID, "fdataf").send_keys(ini_exec_real.strftime("%d/%m/%Y"))
                sleep(1)
                driver.find_element(By.ID, "fhoraf").clear()
                driver.find_element(By.ID, "fhoraf").send_keys(f"{str(fim_exec_real)}:00" if len(str(fim_exec_real)) <= 5 else str(fim_exec_real))
                sleep(2)

                botao_gravar = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Gravar' and contains(@onclick, '4&di')]"))
                )
                botao_gravar.click()
                log(f"Gravado colaborador {colaborador}")
                sleep(2)

                try:
                    WebDriverWait(driver, 3).until(
                        EC.visibility_of_element_located((By.XPATH, "//*[contains(text(), 'Esse Funcionário já possui apontamento nesse período')]"))
                    )
                    log(f"Conflito detectado! Ajustando hora final de {colaborador}")

                    hora_corrigida = (
                        hor_ini_real + timedelta(seconds=1)
                        if isinstance(hor_ini_real, datetime)
                        else datetime.strptime(str(hor_ini_real), "%H:%M:%S") + timedelta(seconds=1)
                    )
                    fhorai_input = driver.find_element(By.ID, "fhorai")
                    fhorai_input.clear()
                    fhorai_input.send_keys(hora_corrigida.strftime("%H:%M:%S"))

                    botao_gravar = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @value='Gravar' and contains(@onclick, '4&di')]"))
                    )
                    botao_gravar.click()
                    sleep(2)

                    log(f"Gravado colaborador {colaborador} após ajuste de horário")

                except TimeoutException:
                    log(f"Lançamento sem conflitos para {colaborador}")
                    sleep(2)

        # Salvar a ordem
        try:
            botao_salvar_ordem = wait.until(
                EC.element_to_be_clickable((By.NAME, "gravaos"))
            )
            botao_salvar_ordem.click()
            log(f"Ordem de serviço {num_os} salva com sucesso")
            sleep(3)
        except Exception as e:
            log(f"Erro ao tentar salvar a OSP {num_os}: {e}")

        # Pegar número da ordem gerada
        num_ordem_input = wait.until(EC.presence_of_element_located((By.ID, "osnum")))
        num_ordem = num_ordem_input.get_attribute("value")
        log(f"OSP gerada com sucesso: {num_ordem}")

        # Atualizar planilha
        OSP = openpyxl.load_workbook('./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx', data_only=False)
        pagina_OSP = OSP['LOGIX X MANUSIS-OSP']
        for i, row in enumerate(pagina_OSP.iter_rows(min_row=2), start=2):
            if row[header.index("num_os")].value == num_os and str(row[header.index("GRUPO")].value).strip() == grupo:
                pagina_OSP.cell(row=i, column=header.index("ORDEM ELECTROLUX") + 1).value = num_ordem
                pagina_OSP.cell(row=i, column=header.index("STATUS MANUSIS") + 1).value = "REALIZADO"
        OSP.save('./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx')

        # Fecha ordem
        checkbox_fecha_os = wait.until(EC.element_to_be_clickable((By.NAME, "fechaos")))
        checkbox_fecha_os.click()
        log("Checkbox 'Fechar OS' marcada.")

        wait.until(EC.text_to_be_present_in_element_value((By.NAME, "gravaos"), "Fechar OS"))

        botao_salvar_ordem = wait.until(EC.element_to_be_clickable((By.NAME, "gravaos")))
        botao_salvar_ordem.click()
        log("Botão 'Fechar OS' clicado com sucesso.")

        driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        log(f"Erro no lançamento do grupo: {e}")

    return lancamentos_realizados


# Abre o site do Manusis
def rodar_automacao(log):
        lancamentos_realizados = 0
        log("Script iniciado")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        log("Acessando Manusis...")
        driver.get('http://electrolux.manusis.com.br/portal/')
        sleep(5)

# - Clicar na opcao FERRAMENTAL
        log("Selecionando sistema Ferramental")

        select_ferramental = driver.find_element(By.ID, "combo_sistemas")

        select_ferramental = Select(select_ferramental)

        select_ferramental.select_by_visible_text("Ferramental")

# - Digitar login
        login = driver.find_element(By.XPATH,'//input[@id="usuario"]')
        sleep(1)
        login.send_keys('HELPTECHSFA')

# - Digitar senha
        senha = driver.find_element(By.XPATH, '//input[@name="senha"]')
        sleep(1)
        senha.send_keys('HELPTECHSFA')

# - Clicar em IR
        wait = WebDriverWait(driver, 10)
        botao_ir = wait.until(EC.element_to_be_clickable((By.NAME, "logar")))

        driver.execute_script("arguments[0].click();", botao_ir)
        sleep(2)

# - Clicar em ORDENS/CARTEIRA DE SERVICOS
        log("Acessando menu de Ordens")
        ordens_menu = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//a[@title="Ordens"]'))
        )

# - Hover (passa o mouse) sobre "Ordens"
        actions = ActionChains(driver)
        actions.move_to_element(ordens_menu).pause(1).perform()  # A pausa ajuda o menu a aparecer

# - Espera a "Carteira de Serviços" aparecer no submenu
        carteira_servicos = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.XPATH, '//a[@title="Carteira de Serviços"]'))
        )

# - Clica com JavaScript para garantir
        driver.execute_script("arguments[0].click();", carteira_servicos)

        sleep(2)

        log("Abrindo nova ordem de serviço")
# - Clicar em ABRIR NOVA ORDEM
        botao_abrir_ordem = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, '//a[contains(@href, "abre_janela_apontaosplan")]'))
        )
        botao_abrir_ordem.click()

# - clicar na janela de lancamento
        driver.switch_to.window(driver.window_handles[-1])

        sleep(5)

# - Extrair informacoes da planilha
        log("Lendo planilha Excel")
        OSP = openpyxl.load_workbook('./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx', data_only=True)
        pagina_OSP = OSP['LOGIX X MANUSIS-OSP']

    

# - Capturar os nomes das colunas (cabeçalho)
        header = [cell.value for cell in next(pagina_OSP.iter_rows(min_row=1, max_row=1))]
        
# - Atribui às variáveis os índices das colunas: status, número da OS, ordem electrolux, início e hora de início da execução
        idx_status = header.index("STATUS MANUSIS")
        idx_num_os = header.index("num_os")
        
    
# - Cria um dicionário para agrupar as ordens de serviço (OS) com status "PENDENTE"
# - Cada OS serve como chave, e seu valor é uma lista com todas as linhas (colaboradores) associadas a ela
        ordens_agrupadas = defaultdict(list)
        for linha in pagina_OSP.iter_rows(min_row=2, values_only=True):
            if linha [idx_status] == "PENDENTE":
                    ordens_agrupadas[linha[idx_num_os]].append(linha)

# - Para cada número de OS, ordena os colaboradores pela data e hora de início de execução real
# - Garante que o colaborador com o início mais cedo seja o primeiro da lista
# - Faz tratamento caso as células estejam como string ou datetime                

        for num_os, linhas in ordens_agrupadas.items():

            def safe_str_strip(value):
                return str(value).strip() if value is not None else ""
            
            grupos_distintos = set(safe_str_strip(linha[header.index("GRUPO")]) for linha in linhas)

            if len(grupos_distintos) > 1:
                print(f"OS: {num_os}, Grupos encontrados: {grupos_distintos}")
                for grupo_atual in grupos_distintos:
                    print(f"Processando grupo: '{grupo_atual}'")
                    linhas_grupo = [linha for linha in linhas if str(linha[header.index("GRUPO")]).strip() == grupo_atual]
                    print(f"Linhas nesse grupo: {len(linhas_grupo)}")
                    lancamentos_realizados += lancar_ordem_para_grupo(driver, log, num_os, linhas_grupo, header)
            else:
                lancamentos_realizados += lancar_ordem_para_grupo(driver, log, num_os, linhas, header)

        return lancamentos_realizados


def enviar_relatorio_manusis():
     
     email_remetente = 'enviaremails05@gmail.com'
     senha_app = 'gfnw gzdi cuqk edkx'

     destinatarios = [
        'gabriel.souza@helptech.ind.br',
        'gabriel.moraes@helptech.ind.br',
        'antonio.silva@helptech.ind.br'  
     ]

     caminho_arquivo = './ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx'
     msg = EmailMessage()
     msg['Subject'] = 'Relatório de Lançamentos no Manusis OSP'
     msg['From'] = email_remetente
     msg['To'] = ', '.join(destinatarios)
     msg.set_content('Prezados,\n\nInformo que o relatório Manusis OSP foi atualizado com novos lançamentos.\n\nAtenciosamente,\nAutomação Python')
     
     with open(caminho_arquivo, 'rb') as f:
        conteudo = f.read()
        nome_arquivo = os.path.basename(caminho_arquivo)
        msg.add_attachment(conteudo, maintype ='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=nome_arquivo)
        try:
          with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
               smtp.login(email_remetente, senha_app)
               smtp.send_message(msg)
               print("E-mail enviado com sucesso!")  
        except Exception as e:
          print(f"Erro ao enviar e-mail: {e}")


if __name__ == "__main__":
    lancamentos_realizados = rodar_automacao(print)
    
    if lancamentos_realizados > 0:
        enviar_relatorio_manusis()
    else:
        print("Nenhum lançamento realizado. E-mail não será enviado.")    


                              
