from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta
from selenium.common.exceptions import TimeoutException, WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import smtplib
from email.message import EmailMessage
import os
import win32com.client as win32

# Atualizar planilha com pywin32 para manter validacao de dados

def atualizar_com_pywin32(lista_atualizacoes, log):
    """
    Recebe uma lista de tuplas (num_os, num_ordem_electrolux) e 
    faz update nas colunas ORDEM ELECTROLUX e STATUS MANUSIS via COM.
    """
# Caminho absoluto para abrir a planilha Excel    
    caminho = os.path.abspath(r'./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx')
    log(f"Abrindo workbook via Excel COM: {caminho!r}")
# Inicializa o Excel via pywin32    
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False

# Abre a planilha e seleciona a aba LOGIX X MANUSIS-OSN
    try:
        wb = excel.Workbooks.Open(caminho)
        sht = wb.Worksheets("LOGIX X MANUSIS-OSN")
    except Exception as e:
        log(f"Não foi possível abrir o Excel: {e}")
        excel.Quit()
        return
# Determina a ultima linha usada para iterar apenas intervalos de dados
    usado = sht.UsedRange
    ultima_linha = usado.Rows.Count
    log(f"Última linha detectada: {ultima_linha}")

# Lê cabeçalho na primeira linha
    header_vals = []
    for col in range(1, sht.UsedRange.Columns.Count + 1):
        header_vals.append(sht.Cells(1, col).Value)

# Localiza as colunas, num_os, ORDEM ELECTROLUX e STATUS MANUSIS
    try:
        idx_num_os    = header_vals.index("num_os") + 1
        idx_ord_elect = header_vals.index("ORDEM ELECTROLUX") + 1
        idx_status    = header_vals.index("STATUS MANUSIS") + 1
    except ValueError as e:
        log(f"Coluna esperada não encontrada no cabeçalho: {e}")
        wb.Close(False)
        excel.Quit()
        return

# Constrói dicionário {num_os → num_ordem_electrolux}
    mapa = {}
    for (osn, ord_eletro) in lista_atualizacoes:
# Tenta converter osn em inteiro, removendo o OSN e mantendo somente os numeros que vem da lista
        try:
            chave = int(osn)  
        except:
            
            digitos = "".join(filter(str.isdigit, str(osn)))
            try:
                chave = int(digitos)
            except:
                log(f"Não foi possível converter '{osn}' em inteiro; pulando.")
                continue
# Converte o ord_eletro em inteiro, ou grava como texto caso falhe
        try:
            valor = int(ord_eletro)
        except:
            try:
                valor = int("".join(filter(str.isdigit, str(ord_eletro))))
            except:
                log(f"Não foi possível converter '{ord_eletro}' em inteiro; será gravado como texto.")
                valor = ord_eletro

        mapa[chave] = valor

    log(f"Mapa de atualizações construído: {mapa}")

# Percorre cada linha da planilha
    atualizou = 0
    for linha in range(2, ultima_linha + 1):
        cell_val = sht.Cells(linha, idx_num_os).Value
        if cell_val is None:
            continue
# Tenta extrair o numero da OS ignorando textos extras que vem da celula do Excel        
        try:
            chave = int(cell_val)
        except:
            
            try:
                chave = int("".join(filter(str.isdigit, str(cell_val))))
            except:
                continue
# Se esse num_os estiver no dicionario, tenta atualizar a celula
        if chave in mapa:
            novo_valor = mapa[chave]
            sht.Cells(linha, idx_ord_elect).Value = novo_valor
            sht.Cells(linha, idx_status).Value = "REALIZADO"
            atualizou += 1
            log(f"Linha {linha}: num_os={chave} → ORDEM ELECTROLUX={novo_valor}, STATUS='REALIZADO'")
# Se houver atualizacao, salva a planilha
    if atualizou > 0:
        log(f"[COM] Salvando workbook ({atualizou} linhas atualizadas)…")
        try:
            wb.Save()
            log("[COM] Workbook salvo com sucesso.")
        except Exception as e:
            log(f"[COM][Erro] Ao salvar: {e}")
    else:
        log("[COM] Nenhuma linha para atualizar; não salvou nada.")
# Fecha o excel
    wb.Close()
    excel.Quit()
    log("[COM] Excel COM finalizado.")


# Função principal que roda a automação no Manusis

def rodar_automacao(log):
  
    lista_para_atualizar = []
    log("INÍCIO DO SCRIPT")
# Tenta inicializar o Chrome    
    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    except WebDriverException as e:
        log(f"Não foi possível iniciar o Chrome: {e}")
        return []

# Timeout para carregar a pagina
    driver.implicitly_wait(5)
    wait = WebDriverWait(driver, 15)

# Entra no site, e faz o login
    try:
        log("Acessando Manusis…")
        driver.get('http://electrolux.manusis.com.br/portal/')
        sleep(3)

        log("Selecionando sistema 'Ferramental'…")
        select_ferramental = Select(driver.find_element(By.ID, "combo_sistemas"))
        select_ferramental.select_by_visible_text("Ferramental")
        sleep(1)

        log("Preenchendo usuário e senha…")
        driver.find_element(By.XPATH, '//input[@id="usuario"]').send_keys('HELPTECHSFA')
        driver.find_element(By.XPATH, '//input[@name="senha"]').send_keys('HELPTECHSFA')
        botao_ir = wait.until(EC.element_to_be_clickable((By.NAME, "logar")))
        botao_ir.click()
        sleep(2)

        log("Entrou no sistema. Navegando para Ordens → Carteira de Serviços…")
        ordens_menu = wait.until(EC.presence_of_element_located((By.XPATH, '//a[@title="Ordens"]')))
        ActionChains(driver).move_to_element(ordens_menu).pause(1).perform()
        carteira = wait.until(EC.visibility_of_element_located((By.XPATH, '//a[@title="Carteira de Serviços"]')))
        driver.execute_script("arguments[0].click();", carteira)
        sleep(2)

    except Exception as e:
        log(f"Ao acessar menu de Ordens: {e}")
        driver.quit()
        return []

# Leitura da planilha
    try:
        OSN_valores = openpyxl.load_workbook('./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx', data_only=True)
        pagina_valores = OSN_valores['LOGIX X MANUSIS-OSN']
        log("Planilha lida com sucesso (data_only=True).")
    except Exception as e:
        log(f"Ao abrir planilha de valores: {e}")
        driver.quit()
        return []

    try:
        OSN = openpyxl.load_workbook('./ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx', data_only=False)
        pagina_OSN = OSN['LOGIX X MANUSIS-OSN']
        log("Planilha carregada (data_only=False) para escrita.")
    except Exception as e:
        log(f"Ao abrir planilha para escrita: {e}")
        driver.quit()
        return []

# Captura cabeçalho para encontrar os indices das colunas necessarias
    header = [cell.value for cell in next(pagina_OSN.iter_rows(min_row=1, max_row=1))]
    try:
        idx_status = header.index("STATUS MANUSIS")
        idx_num_os = header.index("num_os")
        idx_ini_exec_real = header.index("ini_exec_real")
        idx_hor_ini_real = header.index("hor_ini_real")
        idx_ordem_electrolux = header.index("ORDEM ELECTROLUX")
    except ValueError as e:
        log(f"Cabeçalho esperado não encontrado: {e}")
        driver.quit()
        return []

# Monta dicionário {num_os → lista de linhas (tuplas) em que STATUS MANUSIS="PENDENTE"}
    ordens_agrupadas = defaultdict(list)
    for linha in pagina_valores.iter_rows(min_row=2, values_only=True):
        if linha[idx_status] == "PENDENTE":
            ordens_agrupadas[linha[idx_num_os]].append(linha)

    if not ordens_agrupadas:
        log("Não há OS com STATUS 'PENDENTE'. Script encerrando.")
        driver.quit()
        return []

    log(f"Encontrei {len(ordens_agrupadas)} OSNs pendentes.")

# Para cada num_os pendente, abrir janela, preencher dados, salvar, coletar num_ordem_electrolux
    for num_os, linhas in ordens_agrupadas.items():
        log(f"Processando OS {num_os} ──")
        try:
# Volta para a janela principal
            driver.switch_to.window(driver.window_handles[0])
            botao_abrir_ordem = wait.until(
                EC.element_to_be_clickable((By.XPATH, '//a[contains(@href, "abre_janela_apontaosplan")]'))
            )
            botao_abrir_ordem.click()
            wait.until(lambda d: len(d.window_handles) > 1)
            driver.switch_to.window(driver.window_handles[-1])
            sleep(2)
            log(f"Janela de lançamento aberta para OS {num_os}.")
        except Exception as e:
            log(f"Ao abrir janela de lançamento da OS {num_os}: {e}")
            continue

# Ordena colaboradores pela data e hora de início menores
        try:
            def chave_data_hora(linha):
                data = linha[idx_ini_exec_real]
                hora = linha[idx_hor_ini_real]
                if not isinstance(data, datetime):
                    data = datetime.strptime(str(data), "%d/%m/%Y")
                if not isinstance(hora, datetime):
                    tiras = str(hora).strip()
                    if len(tiras.split(":")) == 2:
                        tiras += ":00"
                    hora = datetime.strptime(tiras, "%H:%M:%S")
                return (data, hora)

            linhas_ordenadas = sorted(linhas, key=chave_data_hora)
            base = linhas_ordenadas[0]
            ini_exec_real = chave_data_hora(base)[0]
            hor_ini_real = chave_data_hora(base)[1]
            log(f"[openpyxl] Colaborador-base encontrado: {base[header.index('nom_exec')]} em {ini_exec_real.strftime('%d/%m/%Y')} {hor_ini_real.strftime('%H:%M:%S')}")
        except Exception as e:
            log(f"[Erro] Ao ordenar colaboradores da OS {num_os}: {e}")
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue

# Preenche DATA_PROG, DATA_ABRE e HORA_ABRE
        try:
            data_prog = wait.until(EC.presence_of_element_located((By.ID, "DATA_PROG")))
            data_prog.clear()
            data_prog.send_keys(ini_exec_real.strftime("%d/%m/%Y"))
            sleep(1)
            data_abre = driver.find_element(By.ID, "DATA_ABRE")
            data_abre.clear()
            data_abre.send_keys(ini_exec_real.strftime("%d/%m/%Y"))
            sleep(1)
            hora_abre = driver.find_element(By.ID, "cc[HORA_ABRE]")
            hora_abre.clear()
            hora_abre.send_keys(hor_ini_real.strftime("%H:%M:%S"))
            sleep(1)
            log("Campos de data/hora preenchidos.")
        except Exception as e:
            log(f"Ao preencher data/hora para OS {num_os}: {e}")
            driver.close()
            driver.switch_to.window(driver.window_handles[0])
            continue
            sleep(1)
# Digita PATRIMÔNIO, abre lista, seleciona máquina
        try:
            patrimonio = base[header.index("PATRIMONIO_ELECTROLUX")]
            sleep(1)
            campo_pat = driver.find_element(By.ID, "campo_filtro_cc[MID_MAQUINA]")
            campo_pat.clear()
            sleep(1)
            campo_pat.send_keys(str(patrimonio))
            campo_pat.send_keys(Keys.ENTER)
            sleep(1)
            campo_pat.send_keys(Keys.ENTER)

# Aguarda opção válida aparecer
            wait.until(EC.presence_of_element_located((By.XPATH, '//select[@id="cc[MID_MAQUINA]"]/option[not(@value="0")]')))
            select_maquina = Select(driver.find_element(By.ID, "cc[MID_MAQUINA]"))
            select_maquina.select_by_index(1)
            log(f"Máquina selecionada para patrimônio {patrimonio}.")
        except Exception as e:
            log(f"Ao selecionar máquina para OS {num_os}: {e}")

# Preenche demais campos fixos
        try:
            driver.find_element(By.ID, "cc[SOLICITANTE]").send_keys("MANUFATURA")
            sleep(1)
            driver.find_element(By.ID, "cc[RESPONSAVEL]").send_keys(str(base[header.index("RE")]))
            sleep(1)
            driver.find_element(By.ID, "cc[TIPO_SERVICO]").send_keys("CORRETIVA")
            sleep(1)
            driver.find_element(By.ID, "cc[NATUREZA]").send_keys("3")
            sleep(1)
            driver.find_element(By.ID, "cc[MID_PRIORIDADE]").send_keys("NORMAL")
            sleep(1)
            driver.find_element(By.ID, "cc[TEXTO]").send_keys(str(base[header.index("des_serv_solic")]))
            sleep(1)
            driver.find_element(By.ID, "cc[SOLUCAO_TEXTO]").send_keys(str(base[header.index("des_servico")]))
            log("Campos fixos (solicitante, responsável, etc.) preenchidos.")
        except Exception as e:
            log(f"Ao preencher campos fixos para OS {num_os}: {e}")

            sleep(1)

# Seleciona CAUSA / DEFEITO / SOLUCAO no dropdown
        def selecionar_opcao_parcial(select_element, texto):
            for option in select_element.options:
                if texto.strip().lower() in option.text.strip().lower():
                    select_element.select_by_visible_text(option.text)
                    return option.text
            raise ValueError(f"Opção '{texto}' não encontrada no dropdown.")

        try:
            wait.until(EC.presence_of_element_located((By.ID, "cc[CAUSA]")))
            wait.until(EC.presence_of_element_located((By.ID, "cc[DEFEITO]")))
            wait.until(EC.presence_of_element_located((By.ID, "cc[SOLUCAO]")))
            sleep(1)
            select_falha = Select(driver.find_element(By.ID, "cc[CAUSA]"))
            sleep(1)
            select_defeito = Select(driver.find_element(By.ID, "cc[DEFEITO]"))
            sleep(1)
            select_solucao = Select(driver.find_element(By.ID, "cc[SOLUCAO]"))

            falha_texto = str(base[header.index("FALHA")] or "").strip()
            defeito_texto = str(base[header.index("DEFEITO")] or "").strip()
            solucao_texto = str(base[header.index("SOLUCAO")] or "").strip()

            if falha_texto:
                sel_f = selecionar_opcao_parcial(select_falha, falha_texto)
                log(f"Falha selecionada: {sel_f}")
            if defeito_texto:
                sel_d = selecionar_opcao_parcial(select_defeito, defeito_texto)
                log(f"Defeito selecionado: {sel_d}")
            if solucao_texto:
                sel_s = selecionar_opcao_parcial(select_solucao, solucao_texto)
                log(f"Solução selecionada: {sel_s}")

        except Exception as e:
            log(f"[Selenium][Erro] Ao selecionar Falha/Defeito/Solução: {e}")

# Percorre outros colaboradores da mesma OSN e grava seus apontamentos de horas
        for outra_linha in pagina_valores.iter_rows(min_row=2, values_only=True):
            if outra_linha[idx_status] == "PENDENTE" and outra_linha[header.index("num_os")] == num_os:
                colaborador = outra_linha[header.index("nom_exec")]
                ini_exec = outra_linha[header.index("ini_exec_real")]
                hor_ini = outra_linha[header.index("hor_ini_real")]
                hor_fim = outra_linha[header.index("hor_fim_real")]

                try:
                    driver.find_element(By.ID, "func").send_keys(str(colaborador))
                    sleep(0.5)
                    driver.find_element(By.ID, "fdatai").clear()
                    driver.find_element(By.ID, "fdatai").send_keys(ini_exec.strftime("%d/%m/%Y"))
                    sleep(2)
                    driver.find_element(By.ID, "fhorai").clear()
                    driver.find_element(By.ID, "fhorai").send_keys(f"{str(hor_ini)}:00" if len(str(hor_ini)) <= 5 else str(hor_ini))
                    sleep(2)
                    driver.find_element(By.ID, "fdataf").clear()
                    driver.find_element(By.ID, "fdataf").send_keys(ini_exec.strftime("%d/%m/%Y"))
                    sleep(2)
                    driver.find_element(By.ID, "fhoraf").clear()
                    driver.find_element(By.ID, "fhoraf").send_keys(f"{str(hor_fim)}:00" if len(str(hor_fim)) <= 5 else str(hor_fim))

                    botao_gravar = wait.until(EC.element_to_be_clickable(
                        (By.XPATH, "//input[@type='button' and @value='Gravar' and contains(@onclick, '4&di')]")
                    ))
                    botao_gravar.click()
                    sleep(1)

# Verifica conflito de horário e corrige adicionando 1 segundo
                    try:
                        wait_conf = WebDriverWait(driver, 2)
                        alerta_conf = wait_conf.until(EC.visibility_of_element_located(
                            (By.XPATH, "//*[contains(text(), 'Esse Funcionário já possui apontamento nesse período')]")))
                        log(f"Conflito detectado para {colaborador}; ajustando +1s")
                        hora_corrigida = hor_ini + timedelta(seconds=1) if isinstance(hor_ini, datetime) else datetime.strptime(str(hor_ini), "%H:%M:%S") + timedelta(seconds=1)
                        driver.find_element(By.ID, "fhorai").clear()
                        driver.find_element(By.ID, "fhorai").send_keys(hora_corrigida.strftime("%H:%M:%S"))
                        botao_gravar.click()
                        sleep(1)
                        log(f"Gravou {colaborador} após ajuste de horário.")
                    except TimeoutException:
                        log(f"Lançamento sem conflitos para {colaborador}.")

                except Exception as e:
                    log(f"Ao lançar apontamento de {colaborador}: {e}")
                    sleep(3)

# Salvar OS
        try:
            botao_salvar_ordem = wait.until(
                EC.element_to_be_clickable((By.NAME, "gravaos"))
            )
            botao_salvar_ordem.click()
            log(f"Ordem de serviço {num_os} salva com sucesso.")
            sleep(1)
        except Exception as e:
            log(f"Ao salvar OS {num_os}: {e}")
            sleep(3)

# Recupera número gerado da ORDEM ELECTROLUX
        try:
            num_ordem_input = wait.until(EC.presence_of_element_located((By.ID, "osnum")))
            num_ordem = num_ordem_input.get_attribute("value")
            log(f"OSN gerada: {num_ordem} (para num_os={num_os})")
        except Exception as e:
            log(f"Não conseguiu ler num_ordem para OS {num_os}: {e}")
            num_ordem = None

        if num_ordem:
# Guarda tupla (num_os, num_ordem) para atualizar no Excel COM
            lista_para_atualizar.append((num_os, num_ordem))
            sleep(3)
        
        
# Fecha ordem de servico, clica em "Fechar OS"
            try:

                checkbox_fecha_os = wait.until(EC.element_to_be_clickable((By.NAME, "fechaos")))
                sleep(1)
                checkbox_fecha_os.click()
                log("Checkbox 'Fechar OS' marcada.")

                wait.until(EC.text_to_be_present_in_element_value((By.NAME, "gravaos"), "Fechar OS"))

                botao_salvar_ordem = wait.until(EC.element_to_be_clickable((By.NAME, "gravaos")))
                sleep(1)
                botao_salvar_ordem.click()
                log("Botão 'Fechar OS' clicado com sucesso.")
# Volta para a janela principal
                driver.switch_to.window(driver.window_handles[0])

            except Exception as e:
                log(f"Erro no lançamento do grupo: {e}")


# Fim do loop for num_os
    driver.quit()
    log("FIM DA AUTOMACAO")
    return lista_para_atualizar


# Eniva relatorio por email com a planilha anexada
def enviar_relatorio_manusis(log):
    email_remetente = 'enviaremails05@gmail.com'
    senha_app = 'gfnw gzdi cuqk edkx'

    destinatarios = [
        'gabriel.souza@helptech.ind.br',
        'gabriel.moraes@helptech.ind.br',
        'antonio.silva@helptech.ind.br'
    ]

    caminho_arquivo = './ACOMPANHAMENTO SERVIÇOS MANUSIS.xlsx'
    msg = EmailMessage()
    msg['Subject'] = 'Relatório de Lançamentos no Manusis OSN'
    msg['From'] = email_remetente
    msg['To'] = ', '.join(destinatarios)
    msg.set_content('Prezados,\n\nInformo que o relatório Manusis OSN foi atualizado com novos lançamentos.\n\nAtenciosamente,\nAutomação Python')
# Anexa a planilha ao email
    try:
        with open(caminho_arquivo, 'rb') as f:
            conteudo = f.read()
        nome_arquivo = os.path.basename(caminho_arquivo)
        msg.add_attachment(conteudo,
                           maintype='application',
                           subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           filename=nome_arquivo)
    except Exception as e:
        log(f"[Email][Erro] Ao anexar planilha: {e}")
        return
# Envia o email
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(email_remetente, senha_app)
            smtp.send_message(msg)
            log("[Email] E-mail enviado com sucesso!")
    except Exception as e:
        log(f"[Email][Erro] Ao enviar e-mail: {e}")


# Inicio do script principal
if __name__ == "__main__":
    import sys

# Padronizar formato de horas no print
    def log(texto):
        print(f"{datetime.now().strftime('%H:%M:%S')} - {texto}")
# Roda toa a automacao e retorna uma lista de tuplas
    lista = rodar_automacao(log)

# Verifiaca se houve um lancamento bem-sucedido
    if lista:
        log(f"Vou atualizar {len(lista)} registros no Excel via COM.")
# Atualiza o Excel via pywin32, preenchendo ORDEM ELECTROLUX e STATUS = REALIZADO        
        atualizar_com_pywin32(lista, log)
# Envia email com a planilha atualizada        
        enviar_relatorio_manusis(log)
# Se nao houve leancamento, nao envia o email        
    else:
        log("Nenhum lançamento realizado. Não atualizo planilha nem envio e-mail.")
